# -*- coding: utf-8 -*-
"""server.py — 北斗代理事项闭环看板后端"""
from flask import Flask, jsonify, request, Response, abort, send_file
from pathlib import Path
from datetime import datetime
import json, re, threading, os, io, base64
from urllib.parse import unquote

BASE           = Path(__file__).parent
DATA           = BASE / 'data.json'
LOG_FILE       = BASE / 'oplog.json'
BACKUP         = BASE / 'backup'
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'livox2026')
TEMPLATE_FILE  = BASE / 'template.xlsx'
# GitHub 自动同步配置（在 Render 环境变量中设置）
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN', '')
GITHUB_REPO  = os.environ.get('GITHUB_REPO', 'Dandychen-livox/beidou-kanban')
_lock        = threading.Lock()
_sync_lock   = threading.Lock()   # 防止并发 push

app = Flask(__name__, static_folder=str(BASE), static_url_path='')

# ── 数据读写 ──
def read_data():
    if not DATA.exists(): return []
    return json.loads(DATA.read_text(encoding='utf-8'))

def write_data(rows):
    BACKUP.mkdir(exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    (BACKUP / f'data_{ts}.json').write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')
    DATA.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')
    # 异步同步到 GitHub（不阻塞请求）
    threading.Thread(target=_sync_to_github, daemon=True).start()

def _sync_to_github():
    """将 data.json 推送到 GitHub，防止 Render 重启丢数据"""
    if not GITHUB_TOKEN:
        return  # 未配置 Token 则跳过
    try:
        import urllib.request, urllib.error
        with _sync_lock:
            content_raw = DATA.read_bytes()
            content_b64 = base64.b64encode(content_raw).decode()
            api_url = f'https://api.github.com/repos/{GITHUB_REPO}/contents/data.json'
            # 先获取当前 SHA
            req = urllib.request.Request(api_url)
            req.add_header('Authorization', f'token {GITHUB_TOKEN}')
            req.add_header('Accept', 'application/vnd.github.v3+json')
            try:
                with urllib.request.urlopen(req, timeout=10) as r:
                    info = json.loads(r.read())
                sha = info.get('sha', '')
            except Exception:
                sha = ''
            # 推送
            body = json.dumps({
                'message': f'Auto-sync data.json {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
                'content': content_b64,
                'sha': sha
            }).encode()
            req2 = urllib.request.Request(api_url, data=body, method='PUT')
            req2.add_header('Authorization', f'token {GITHUB_TOKEN}')
            req2.add_header('Content-Type', 'application/json')
            req2.add_header('Accept', 'application/vnd.github.v3+json')
            with urllib.request.urlopen(req2, timeout=15) as r:
                pass  # 成功
    except Exception:
        pass  # 同步失败不影响主流程

# ── 操作日志 ──
def read_log():
    if not LOG_FILE.exists(): return []
    try:
        return json.loads(LOG_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []

def write_log(entry):
    logs = read_log()
    logs.append(entry)
    logs = logs[-500:]
    LOG_FILE.write_text(json.dumps(logs, ensure_ascii=False, indent=2), encoding='utf-8')

def add_log(operator, action, item_id=None, item_name=None, detail=None):
    try:
        write_log({
            'time':      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'operator':  operator,
            'action':    action,
            'item_id':   item_id,
            'item_name': item_name,
            'detail':    detail or ''
        })
    except Exception:
        pass

def get_caller(req, body=None):
    """安全获取调用者姓名，处理URL编码的中文"""
    raw = req.headers.get('X-Person', '')
    if not raw and body:
        raw = body.get('caller', '')
    try:
        return unquote(raw).strip() or '匿名'
    except Exception:
        return raw.strip() or '匿名'

def is_admin(req):
    return req.headers.get('X-Admin-Token') == ADMIN_PASSWORD

# ── 路由 ──

@app.route('/')
def index():
    content = (BASE / 'kanban.html').read_text(encoding='utf-8')
    return Response(content, mimetype='text/html; charset=utf-8')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

@app.route('/api/data')
def api_data():
    return jsonify(read_data())

@app.route('/api/auth', methods=['POST'])
def api_auth():
    body = request.get_json(force=True) or {}
    if body.get('password') == ADMIN_PASSWORD:
        add_log('管理员', '登录')
        return jsonify({'ok': True, 'token': ADMIN_PASSWORD})
    return jsonify({'ok': False, 'msg': '密码错误'}), 401

@app.route('/api/log')
def api_log():
    if not is_admin(request): abort(403)
    logs = read_log()
    logs.reverse()
    return jsonify(logs)

@app.route('/api/public_update/<int:row_id>', methods=['POST'])
def api_public_update(row_id):
    body   = request.get_json(force=True) or {}
    caller = get_caller(request, body)
    allowed = {k: v for k, v in body.items() if k in ('person', 'progress', 'submit_url', 'status')}
    if not allowed: abort(400)
    with _lock:
        rows = read_data()
        idx  = next((i for i, r in enumerate(rows) if str(r.get('id')) == str(row_id)), None)
        if idx is None: abort(404)
        row = rows[idx]
        row.update(allowed)
        row['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        row['updated_by'] = caller
        rows[idx] = row
        write_data(rows)
    parts = []
    if 'progress'   in allowed: parts.append('进展:' + allowed['progress'][:30])
    if 'person'     in allowed: parts.append('责任人→' + allowed['person'])
    if 'status'     in allowed: parts.append('状态→' + allowed['status'])
    if 'submit_url' in allowed: parts.append('提交物:' + allowed['submit_url'][:30])
    add_log(caller, '公开填写', row_id, row.get('item', '')[:20], '；'.join(parts))
    return jsonify({'ok': True, 'row': row})

@app.route('/api/update/<int:row_id>', methods=['POST'])
def api_update(row_id):
    body   = request.get_json(force=True) or {}
    admin  = is_admin(request)
    caller = get_caller(request)
    with _lock:
        rows = read_data()
        idx  = next((i for i, r in enumerate(rows) if str(r.get('id')) == str(row_id)), None)
        if idx is None: abort(404)
        row = rows[idx]
        if not admin:
            ps = [p.strip() for p in re.split(r'[&/、,，]', row.get('person', '')) if p.strip()]
            if caller not in ps: abort(403)
            body = {k: v for k, v in body.items() if k in ('progress', 'status', 'livox_confirm')}
        row.update(body)
        row['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        row['updated_by'] = caller or ('管理员' if admin else '?')
        rows[idx] = row
        write_data(rows)
    parts = []
    if 'progress'      in body: parts.append('进展:' + body['progress'][:30])
    if 'status'        in body: parts.append('状态→' + body['status'])
    if 'livox_confirm' in body: parts.append('Livox确认→' + body['livox_confirm'])
    if 'person'        in body: parts.append('责任人→' + body['person'])
    add_log('管理员' if admin else caller, '编辑事项', row_id, row.get('item', '')[:20], '；'.join(parts))
    return jsonify({'ok': True, 'row': row})

@app.route('/api/add', methods=['POST'])
def api_add():
    if not is_admin(request): abort(403)
    body = request.get_json(force=True) or {}
    with _lock:
        rows   = read_data()
        new_id = max((r.get('id', 0) for r in rows), default=0) + 1
        row = {
            'id':           new_id,
            'date':         datetime.now().strftime('%Y-%m-%d'),
            'item':         body.get('item', ''),
            'submit':       body.get('submit', ''),
            'ddl':          body.get('ddl', ''),
            'person':       body.get('person', ''),
            'livox':        body.get('livox', 'Dandy'),
            'progress':     '',
            'submit_url':   body.get('submit_url', ''),
            'status':       '未完成',
            'livox_confirm':'未完成',
            'updated_at':   datetime.now().strftime('%Y-%m-%d %H:%M'),
            'updated_by':   '管理员'
        }
        rows.append(row)
        write_data(rows)
    add_log('管理员', '新增事项', new_id, row['item'][:20])
    return jsonify({'ok': True, 'row': row})

@app.route('/api/delete/<int:row_id>', methods=['DELETE'])
def api_delete(row_id):
    if not is_admin(request): abort(403)
    with _lock:
        rows = read_data()
        idx  = next((i for i, r in enumerate(rows) if str(r.get('id')) == str(row_id)), None)
        if idx is None: abort(404)
        item_name = rows[idx].get('item', '')[:20]
        rows.pop(idx)
        write_data(rows)
    add_log('管理员', '删除事项', row_id, item_name)
    return jsonify({'ok': True})

@app.route('/api/template')
def api_template():
    if TEMPLATE_FILE.exists():
        return send_file(str(TEMPLATE_FILE), as_attachment=True,
                         download_name='事项明细模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '事项明细'
        headers = ['事项名称*', '提交内容/要求', 'DDL', '责任人', 'LIVOX对接人', '进展', '提交物链接', '闭环状态']
        notes   = ['必填', '提交要求', '如：2026-08-15或每月更新', '如：赵云飞', '如：Dandy', '进展说明', 'https://...', '未完成/完成/挂起']
        for i, (h, n) in enumerate(zip(headers, notes), 1):
            ws.cell(1, i, h)
            ws.cell(2, i, n)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(buf.read(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment;filename=template.xlsx'})
    except Exception as e:
        abort(500, str(e))

@app.route('/api/batch_add', methods=['POST'])
def api_batch_add():
    if not is_admin(request): abort(403)
    parse_only = request.headers.get('X-Parse-Only', '') == '1'
    if request.content_type and 'application/json' in request.content_type:
        body = request.get_json(force=True) or {}
        if isinstance(body, list):
            return _do_batch(body)
        b64 = body.get('file_base64', '')
        po  = body.get('parse_only', parse_only)
        if b64:
            try:
                items = _parse_excel(io.BytesIO(base64.b64decode(b64)))
                if po: return jsonify({'ok': True, 'rows': items})
                return _do_batch(items)
            except Exception as e:
                return jsonify({'ok': False, 'msg': '解析失败：' + str(e)}), 400
        abort(400)
    f = request.files.get('file')
    if not f: abort(400)
    try:
        items = _parse_excel(f.stream)
        if parse_only: return jsonify({'ok': True, 'rows': items})
        return _do_batch(items)
    except Exception as e:
        return jsonify({'ok': False, 'msg': '解析失败：' + str(e)}), 400

def _parse_excel(stream):
    import openpyxl
    wb  = openpyxl.load_workbook(stream, data_only=True)
    ws  = wb.active
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or '').strip()
        if v: col_map[v] = c
    FIELD_MAP = {
        '事项名称*': 'item', '事项名称': 'item',
        '提交内容/要求': 'submit', '提交内容': 'submit',
        'DDL': 'ddl', '截止日期': 'ddl',
        '责任人': 'person',
        'LIVOX对接人': 'livox', 'Livox对接人': 'livox',
        '进展': 'progress',
        '提交物链接': 'submit_url', '提交物': 'submit_url',
        '闭环状态': 'status', '状态': 'status',
    }
    items = []
    for r in range(3, ws.max_row + 1):
        row_vals = {k: ws.cell(r, v).value for k, v in col_map.items()}
        if not any(row_vals.values()): continue
        item = {}
        for col_name, field in FIELD_MAP.items():
            if col_name in row_vals and row_vals[col_name] is not None:
                item[field] = str(row_vals[col_name]).strip()
        if not item.get('item'): continue
        items.append(item)
    return items

def _do_batch(items):
    if not items: return jsonify({'ok': False, 'msg': '没有可导入的数据'})
    added = []
    with _lock:
        rows   = read_data()
        cur_id = max((r.get('id', 0) for r in rows), default=0)
        for item in items:
            cur_id += 1
            status = item.get('status', '未完成')
            if status not in ('完成', '未完成', '挂起'): status = '未完成'
            row = {
                'id': cur_id, 'date': datetime.now().strftime('%Y-%m-%d'),
                'item': item.get('item', ''), 'submit': item.get('submit', ''),
                'ddl':  item.get('ddl', ''),  'person': item.get('person', ''),
                'livox': item.get('livox', 'Dandy'), 'progress': item.get('progress', ''),
                'submit_url': item.get('submit_url', ''), 'status': status,
                'livox_confirm': '未完成',
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                'updated_by': '管理员(批量导入)',
            }
            rows.append(row)
            added.append(row)
        write_data(rows)
    add_log('管理员', '批量导入', None, None, '导入' + str(len(added)) + '条事项')
    return jsonify({'ok': True, 'added': len(added), 'rows': added})

if __name__ == '__main__':
    print('=' * 50)
    print('北斗代理事项闭环看板 已启动')
    print('本机：http://127.0.0.1:8080')
    print('内网：http://192.168.255.10:8080')
    print('=' * 50)
    app.run(host='0.0.0.0', port=8080, debug=False)
